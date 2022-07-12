#!/usr/bin/env python
# -*- coding:utf-8 -*-
import requests
import re
import os
from bs4 import BeautifulSoup
from openpyxl import Workbook

if __name__ == "__main__":
    # 创建文件夹
    if not os.path.exists('./banners'):
        os.mkdir('./banners')

    # 获取整个页面
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36'
    }
    url = 'http://www.cumt.edu.cn/'
    page_text = requests.get(url=url, headers=headers).content.decode('utf-8')

    # 查找到图片链接的数据
    soup = BeautifulSoup(page_text, 'lxml')
    w2imgJsons = soup.select('#wp_news_w2 > script')[0].text

    # 正则匹配出图片链接地址
    ex = 'src:"(.*?)"'
    img_url_list = re.findall(ex, w2imgJsons, re.S)

    # 打印到控制台、保存到本地、保存到excel
    wb = Workbook()
    ws = wb.create_sheet('banners', 0)
    for i, src in enumerate(img_url_list):
        img_url = 'http://www.cumt.edu.cn/' + src
        print(img_url)
        ws['A'+str(i+1)] = img_url
        img_data = requests.get(url=img_url, headers=headers).content
        img_name = img_url.split('/')[-1]
        img_path = './banners/' + img_name
        with open(img_path, 'wb') as fp:
            fp.write(img_data)
            print(img_name, '下载成功！')
    wb.save('Banners.xlsx')
    wb.close()
